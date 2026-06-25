"""
FUR CPR Automation System - Backend
Dragon Studios for FUR Store (furstores.myshopify.com)
"""

import os, re, time, json, threading
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import pdfplumber
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

app = Flask(__name__)
CORS(app)

# ── In-memory job store (replace with Redis/DB for production) ──
jobs = {}

SHOPIFY_STORE = "furstores.myshopify.com"
API_VERSION = "2024-10"

def shopify_graphql(token, query, variables=None):
    url = f"https://{SHOPIFY_STORE}/admin/api/{API_VERSION}/graphql.json"
    try:
        resp = requests.post(url,
            headers={
                "Content-Type": "application/json",
                "X-Shopify-Access-Token": token
            },
            json={"query": query, "variables": variables or {}},
            timeout=30
        )
        result = resp.json()
        # Always return a dict
        if not isinstance(result, dict):
            return {"errors": [{"message": f"Unexpected response type: {type(result)}"}]}
        return result
    except requests.exceptions.Timeout:
        return {"errors": [{"message": "Request timed out. Shopify API slow hai — dobara try karo."}]}
    except requests.exceptions.ConnectionError:
        return {"errors": [{"message": "Shopify se connect nahi ho saka. Token check karo."}]}
    except Exception as e:
        return {"errors": [{"message": str(e)}]}


def parse_cpr_pdf(file_bytes):
    """Parse PostEx CPR PDF and extract all tracking numbers with status/COD."""
    tracking_re = re.compile(r'\b(\d{14})\b')
    cod_re      = re.compile(r'(\d{1,3}(?:,\d{3})*\.\d{2})')
    cpr_no_re   = re.compile(r'CPR[- ]?([\w\d]+)')
    date_re     = re.compile(r'(\d{2}/\d{2}/\d{4})')

    delivered, returned = [], []
    cpr_number, cpr_date = "", ""
    summary = {}

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""

            # Extract CPR number and date from first page
            if not cpr_number:
                m = cpr_no_re.search(text)
                if m:
                    cpr_number = m.group(0).strip()
            if not cpr_date:
                dates = date_re.findall(text)
                if dates:
                    cpr_date = dates[0]

            # Extract summary counts
            if "Delivered" in text and not summary:
                del_m = re.search(r'Delivered\s+(\d+)\s+([\d,]+\.\d+)', text)
                ret_m = re.search(r'Returned\s+(\d+)\s+([\d,]+\.\d+)', text)
                net_m = re.search(r'Net Payable.*?([\d,]+\.\d+)', text)
                if del_m:
                    summary['delivered_count'] = int(del_m.group(1))
                    summary['cod_collected']   = float(del_m.group(2).replace(',',''))
                if ret_m:
                    summary['returned_count'] = int(ret_m.group(1))
                if net_m:
                    summary['net_payable'] = float(net_m.group(1).replace(',',''))

            # Extract orders line by line
            lines = text.split('\n')
            for j, line in enumerate(lines):
                t_match = tracking_re.search(line)
                if not t_match:
                    continue
                tracking = t_match.group(1)
                next_line = lines[j+1].strip() if j+1 < len(lines) else ""
                combined  = line + " " + next_line

                status = (
                    "delivered" if "Delivered" in combined else
                    "returned"  if "Return"    in combined else
                    "unknown"
                )
                cods = cod_re.findall(combined)
                cod  = float(cods[0].replace(',','')) if cods else 0.0

                entry = {"tracking": tracking, "cod": cod, "status": status}
                if status == "delivered":
                    delivered.append(entry)
                elif status == "returned":
                    returned.append(entry)

    return {
        "cpr_number": cpr_number,
        "cpr_date":   cpr_date,
        "delivered":  delivered,
        "returned":   returned,
        "summary":    summary
    }


def fetch_all_shopify_orders(token, job_id):
    """Paginate through Shopify orders and index by tracking number."""
    all_orders = {}
    cursor = None
    page = 0

    query = """
    query GetOrders($after: String) {
        orders(first: 50, after: $after, sortKey: CREATED_AT,
               query: "created_at:>=2026-05-01 created_at:<=2026-07-31") {
            pageInfo { hasNextPage endCursor }
            edges { node {
                id name displayFinancialStatus cancelledAt
                fulfillments(first: 1) { trackingInfo { number } }
            }}
        }
    }"""

    while True:
        page += 1
        variables = {"after": cursor} if cursor else {}
        data = shopify_graphql(token, query, variables)

        # Check for API errors
        if not isinstance(data, dict):
            return None, f"Shopify ne unexpected response diya: {data}"

        if data.get("errors"):
            err_msg = data["errors"][0].get("message", "Unknown Shopify error")
            # Common error: invalid token
            if "401" in str(err_msg) or "Unauthorized" in str(err_msg) or "Invalid API key" in str(err_msg):
                return None, "Shopify token invalid hai. Nayi token banao aur dobara try karo."
            return None, f"Shopify API error: {err_msg}"

        gql_data = data.get("data")
        if not gql_data:
            # Could be auth error without proper error field
            ext = data.get("extensions", {})
            return None, f"Shopify se data nahi aaya. Token check karo. Response: {str(data)[:200]}"

        orders_obj = gql_data.get("orders", {})
        if not orders_obj:
            return None, "Orders field missing in response"

        for edge in orders_obj.get("edges", []):
            node = edge.get("node", {})
            fulfillments = node.get("fulfillments") or []
            if fulfillments:
                tracking_list = fulfillments[0].get("trackingInfo") or []
                if tracking_list:
                    tracking_num = tracking_list[0].get("number", "")
                    if tracking_num:
                        all_orders[tracking_num] = node

        jobs[job_id]["progress"] = min(40, page * 6)
        jobs[job_id]["log"].append({
            "type": "info",
            "msg":  f"Page {page}: {len(all_orders)} orders indexed so far"
        })

        page_info = orders_obj.get("pageInfo", {})
        if not page_info.get("hasNextPage"):
            break
        cursor = page_info.get("endCursor")
        time.sleep(0.25)

    return all_orders, None


def run_automation(job_id, token, cpr_data):
    """Main automation runner — runs in background thread."""
    job = jobs[job_id]
    job["status"] = "running"
    job["log"] = []

    def log(msg, t="info", order=None):
        job["log"].append({"type": t, "msg": msg, "order": order})

    try:
        log("═══ PHASE 1: Fetching Shopify orders ═══")
        all_orders, err = fetch_all_shopify_orders(token, job_id)
        if err:
            job["status"] = "error"
            job["error"]  = err
            log(f"Shopify error: {err}", "error")
            return

        log(f"Total orders indexed: {len(all_orders)}", "success")

        # ── PHASE 2: Verify returns ──
        log("═══ PHASE 2: Verifying return orders ═══")
        returns_voided     = []
        returns_not_canceled = []
        returns_not_found  = []

        for ret in cpr_data["returned"]:
            o = all_orders.get(ret["tracking"])
            if not o:
                returns_not_found.append(ret)
                log(f"Not in Shopify: {ret['tracking']}", "warn")
            elif o["displayFinancialStatus"] == "VOIDED" or o["cancelledAt"]:
                returns_voided.append({**ret, "orderName": o["name"], "orderId": o["id"]})
                log(f"Canceled ✓  {ret['tracking']} → #{o['name']}", "success", o["name"])
            else:
                returns_not_canceled.append({
                    **ret, "orderName": o["name"],
                    "orderId": o["id"], "status": o["displayFinancialStatus"]
                })
                log(f"NOT CANCELED: {ret['tracking']} → #{o['name']} ({o['displayFinancialStatus']})", "error", o["name"])

        job["progress"] = 45

        # ── PHASE 3: Mark delivered as PAID ──
        log("═══ PHASE 3: Marking delivered orders as PAID ═══")
        paid_orders    = []
        skipped_orders = []
        error_orders   = []
        not_found      = []

        mutation = """
        mutation MarkPaid($id: ID!) {
            orderMarkAsPaid(input: { id: $id }) {
                order { name displayFinancialStatus }
                userErrors { field message }
            }
        }"""

        delivered = cpr_data["delivered"]
        total = len(delivered)

        for i, del_order in enumerate(delivered):
            o = all_orders.get(del_order["tracking"])
            progress = 45 + int((i / total) * 50)
            job["progress"] = progress

            if not o:
                not_found.append(del_order)
                log(f"Not found: {del_order['tracking']}", "warn")
                continue

            name = o["name"]
            status = o["displayFinancialStatus"]

            if del_order["cod"] == 0:
                skipped_orders.append({**del_order, "orderName": name, "reason": "COD=0"})
                log(f"COD=0 skip: #{name}", "skip", name)
                continue

            if status == "PAID":
                skipped_orders.append({**del_order, "orderName": name, "reason": "Already PAID"})
                log(f"Already paid: #{name}", "skip", name)
                continue

            if status == "VOIDED" or o["cancelledAt"]:
                skipped_orders.append({**del_order, "orderName": name, "reason": "Canceled"})
                log(f"Canceled skip: #{name}", "skip", name)
                continue

            # Execute mark-as-paid
            result = shopify_graphql(token, mutation, {"id": o["id"]})
            errs   = result.get("data", {}).get("orderMarkAsPaid", {}).get("userErrors", [])

            if result.get("errors") or errs:
                err_msg = (result.get("errors") or [{}])[0].get("message") or (errs[0].get("message") if errs else "Unknown")
                error_orders.append({**del_order, "orderName": name, "error": err_msg})
                log(f"FAILED #{name}: {err_msg}", "error", name)
            else:
                paid_orders.append({**del_order, "orderName": name, "orderId": o["id"]})
                log(f"PAID ✓  {del_order['tracking']} → #{name} · Rs {del_order['cod']:,.0f}", "success", name)

            time.sleep(0.2)

        job["progress"] = 100
        job["status"]   = "done"
        job["results"]  = {
            "cpr_number":          cpr_data["cpr_number"],
            "cpr_date":            cpr_data["cpr_date"],
            "paid":                paid_orders,
            "skipped":             skipped_orders,
            "errors":              error_orders,
            "not_found":           not_found,
            "returns_voided":      returns_voided,
            "returns_not_canceled": returns_not_canceled,
            "returns_not_found":   returns_not_found,
        }

        paid_total = sum(o["cod"] for o in paid_orders)
        log(f"COMPLETE — Paid: {len(paid_orders)} (Rs {paid_total:,.0f}) | Skipped: {len(skipped_orders)} | Errors: {len(error_orders)}", "success")
        log(f"Returns: {len(returns_voided)}/{len(cpr_data['returned'])} voided | Action needed: {len(returns_not_canceled)}", "info")

    except Exception as e:
        job["status"] = "error"
        job["error"]  = str(e)
        log(f"Unexpected error: {e}", "error")


def generate_excel_report(results):
    """Generate styled Excel report."""
    wb = openpyxl.Workbook()

    # Colors
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    amber_fill  = PatternFill("solid", fgColor="FFEB9C")
    header_fill = PatternFill("solid", fgColor="1C3557")
    header_font = Font(color="FFFFFF", bold=True)
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center")

    def make_sheet(ws, title, headers, rows, fill=None):
        ws.title = title
        ws.append(headers)
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = center
        for row in rows:
            ws.append(row)
            if fill:
                for cell in ws[ws.max_row]:
                    cell.fill = fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    # Summary sheet
    ws0 = wb.active
    ws0.title = "Summary"
    ws0["A1"] = "FUR CPR Automation Report"
    ws0["A1"].font = Font(size=14, bold=True, color="1C3557")
    ws0["A2"] = f"CPR: {results.get('cpr_number','')}  |  Date: {results.get('cpr_date','')}  |  Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    summary_data = [
        ["", ""],
        ["METRIC", "VALUE"],
        ["Delivered Orders Processed", len(results["paid"]) + len(results["skipped"]) + len(results["errors"]) + len(results["not_found"])],
        ["Marked as PAID",             len(results["paid"])],
        ["Total COD Collected",        f"Rs {sum(o['cod'] for o in results['paid']):,.2f}"],
        ["Skipped (Already paid/COD=0)", len(results["skipped"])],
        ["Failed/Errors",              len(results["errors"])],
        ["Not Found in Shopify",       len(results["not_found"])],
        ["", ""],
        ["Return Orders Total",        len(results["returns_voided"]) + len(results["returns_not_canceled"]) + len(results["returns_not_found"])],
        ["Returns Verified Voided",    len(results["returns_voided"])],
        ["Returns NOT Canceled ⚠",    len(results["returns_not_canceled"])],
    ]
    for row in summary_data:
        ws0.append(row)
    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 25

    # Paid Orders
    paid_rows = [[o["tracking"], f"#{o['orderName']}", f"Rs {o['cod']:,.2f}"] for o in results["paid"]]
    make_sheet(wb.create_sheet(), "✅ Paid Orders", ["Tracking Number", "Order", "COD Amount"], paid_rows, green_fill)

    # Skipped
    skip_rows = [[o["tracking"], f"#{o.get('orderName','')}", f"Rs {o['cod']:,.2f}", o.get("reason","")] for o in results["skipped"]]
    make_sheet(wb.create_sheet(), "⚠ Skipped", ["Tracking Number", "Order", "COD Amount", "Reason"], skip_rows, amber_fill)

    # Errors
    if results["errors"]:
        err_rows = [[o["tracking"], f"#{o.get('orderName','')}", f"Rs {o['cod']:,.2f}", o.get("error","")] for o in results["errors"]]
        make_sheet(wb.create_sheet(), "❌ Errors", ["Tracking Number", "Order", "COD", "Error"], err_rows, red_fill)

    # Returns Voided
    ret_rows = [[o["tracking"], f"#{o.get('orderName','')}", o.get("status","VOIDED")] for o in results["returns_voided"]]
    make_sheet(wb.create_sheet(), "✅ Returns Voided", ["Tracking Number", "Order", "Status"], ret_rows, green_fill)

    # Returns Action Needed
    if results["returns_not_canceled"]:
        act_rows = [[o["tracking"], f"#{o.get('orderName','')}", o.get("status","")] for o in results["returns_not_canceled"]]
        make_sheet(wb.create_sheet(), "🚨 Returns-Action Needed", ["Tracking Number", "Order", "Status"], act_rows, red_fill)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── API Routes ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload", methods=["POST"])
def upload_cpr():
    """Parse CPR PDF and return extracted data for preview."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Only PDF files supported"}), 400

    try:
        data = parse_cpr_pdf(f.read())
        return jsonify({
            "ok":          True,
            "cpr_number":  data["cpr_number"],
            "cpr_date":    data["cpr_date"],
            "delivered":   len(data["delivered"]),
            "returned":    len(data["returned"]),
            "summary":     data["summary"],
            "preview":     {
                "delivered": data["delivered"][:5],
                "returned":  data["returned"][:5],
            },
            "_data":       data,   # kept server-side via job
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/run", methods=["POST"])
def run_job():
    """Start automation job."""
    try:
        body = request.get_json(force=True, silent=True) or {}
    except Exception:
        body = {}

    token    = (body.get("token") or "").strip()
    cpr_data = body.get("cpr_data")

    if not token:
        return jsonify({"error": "Shopify API token required"}), 400
    if not cpr_data:
        return jsonify({"error": "CPR data missing — upload a file first"}), 400

    # Validate cpr_data structure
    if not isinstance(cpr_data, dict):
        return jsonify({"error": f"Invalid CPR data format: {type(cpr_data).__name__}"}), 400
    if "delivered" not in cpr_data or "returned" not in cpr_data:
        return jsonify({"error": "CPR data incomplete — please re-upload the PDF"}), 400

    job_id = f"job_{int(time.time()*1000)}"
    jobs[job_id] = {
        "id":       job_id,
        "status":   "queued",
        "progress": 0,
        "log":      [],
        "results":  None,
        "error":    None,
        "started":  datetime.now().isoformat(),
    }

    t = threading.Thread(target=run_automation, args=(job_id, token, cpr_data), daemon=True)
    t.start()

    return jsonify({"job_id": job_id})


@app.route("/api/status/<job_id>")
def job_status(job_id):
    """Poll job status and logs."""
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    resp = {
        "status":   job["status"],
        "progress": job["progress"],
        "log":      job["log"],
        "error":    job.get("error"),
    }

    if job["status"] == "done" and job["results"]:
        r = job["results"]
        resp["summary"] = {
            "paid":             len(r["paid"]),
            "paid_total":       sum(o["cod"] for o in r["paid"]),
            "skipped":          len(r["skipped"]),
            "errors":           len(r["errors"]),
            "not_found":        len(r["not_found"]),
            "returns_voided":   len(r["returns_voided"]),
            "returns_action":   len(r["returns_not_canceled"]),
        }
        if r["returns_not_canceled"]:
            resp["returns_not_canceled"] = r["returns_not_canceled"]

    return jsonify(resp)


@app.route("/api/report/<job_id>")
def download_report(job_id):
    """Download Excel report for completed job."""
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "Job not done"}), 404

    buf = generate_excel_report(job["results"])
    cpr = job["results"].get("cpr_number", "CPR")
    date_str = datetime.now().strftime("%Y-%m-%d")

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"FUR-{cpr}-Report-{date_str}.xlsx"
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n🚀 FUR CPR Automation running on http://localhost:{port}\n")
    app.run(debug=False, host="0.0.0.0", port=port)
